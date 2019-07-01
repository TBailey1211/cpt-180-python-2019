#Program Name: editSalesInfo.py
#Program Description: Edit an excel spreadsheet using Python
#Programmer's Name: Taylor Bailey

import openpyxl
wb = openpyxl.load_workbook('salsa.xlsx')

ws = wb['Projected Sales']

ws['B5'] = ws['B5'].value * 1.1
ws['B6'] = ws['B6'].value * 1.1
ws['B7'] = ws['B7'].value * 1.1
ws['B8'] = ws['B8'].value * 1.1
ws['B9'] = ws['B9'].value * 1.1

wb.save('salsa2.xlsx')

wb = openpyxl.load_workbook('salsa2.xlsx')
ws = wb['Projected Sales']

for i in range(15, 27):
    refObj = openpyxl.chart.Reference(ws, min_col=3, min_row=5, max_col=3, max_row=9)
    seriesObj = openpyxl.chart.Series(refObj, title='Salsa Prices')
    chartObj = openpyxl.chart.BarChart()
    chartObj.append(seriesObj)
    catRefObj = openpyxl.chart.Reference(ws, min_col=1, min_row=5, max_col=1, max_row=9)
    chartObj.set_categories(catRefObj)

    ws.add_chart(chartObj,'A16')
    wb.save('salsa2.xlsx')
