#Program Name: getSalesInfo.py
#Program Description: Get sales information from an excel spreadsheet
#Programmer's Name: Taylor Bailey

import openpyxl, pprint
wb = openpyxl.load_workbook('salsa.xlsx', data_only=True)
ws = wb['Projected Sales']
sales = {}

for x in range(5,10):
    sales[ws['A' + str(x)].value] = ws['F' + str(x)].value

pprint.pprint(sales)

for x in sales:
    x = sum(sales.values())

print("The total is $" + '{:10,.2f}'.format(x))
