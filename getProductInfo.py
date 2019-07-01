#Program Name: getProductInfo.py
#Program Description: Print out product info from an excel spreadsheet
#Programmer's Name: Taylor Bailey

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

wb = openpyxl.load_workbook('salsa.xlsx')
print(wb.sheetnames)

ws = wb['Product List']
for row in range(4,9):
    print(ws.cell(row = row, column = 2).value)

printRow = ws.max_row
print(row)
printColumn = get_column_letter(ws.max_column)
print(printColumn)
