import openpyxl, pprint
print('Opening workbook...')
wb = openpyxl.load_workbook('censuspopdata.xlsx')
ws = wb['Population by Census Tract']
countyData = {}

print('Reading rows...')
for row in range(2, ws.max_row + 1):
    state = ws['B' + str(row)].value
    county = ws['C' + str(row)].value
    pop = ws['D' + str(row)].value

countyData.setdefault(state, {})
countyData[state].setdefault(county, {'tracts': 0, 'pop': 0})
countyData[state][county]['tracts'] += 1
countyData[state][county]['pop'] += int(pop)

print(countyData)
